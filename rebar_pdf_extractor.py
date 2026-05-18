"""
Rebar Schedule PDF Extractor
=============================
Author  : Divyanshu Raghuwanshi
GitHub  : https://github.com/divyanshu964
Purpose : Reads structural rebar schedule PDFs, extracts bar details,
          calculates weights using standard formula (D² / 162),
          and exports results to a clean Excel file.

Supports:
  - Native vector PDFs (text-selectable) via pdfplumber
  - Scanned image PDFs via Tesseract OCR + OpenCV (offline, no API needed)

Usage:
  python rebar_pdf_extractor.py --pdf your_rebar_schedule.pdf
  python rebar_pdf_extractor.py --demo        # generates a sample PDF and runs extraction on it

Requirements:
  pip install pdfplumber pymupdf pandas openpyxl fpdf2 pillow pytesseract opencv-python-headless

  For scanned PDFs, also install Tesseract OCR engine:
    Mac   : brew install tesseract
    Linux : sudo apt install tesseract-ocr
    Windows: https://github.com/UB-Mannheim/tesseract/wiki
"""

import argparse
import re
import sys
import os
import warnings
warnings.filterwarnings("ignore")

# ── third-party imports (installed via pip) ──────────────────────────────────
try:
    import pdfplumber
except ImportError:
    sys.exit("[ERROR] pdfplumber not found. Run: pip install pdfplumber")

try:
    import pandas as pd
except ImportError:
    sys.exit("[ERROR] pandas not found. Run: pip install pandas openpyxl")

try:
    from fpdf import FPDF
except ImportError:
    sys.exit("[ERROR] fpdf2 not found. Run: pip install fpdf2")


# ─────────────────────────────────────────────────────────────────────────────
# 1.  WEIGHT FORMULA  (Indian Standard / universal BBS formula)
#     Unit weight (kg/m) = D² / 162   where D = diameter in mm
# ─────────────────────────────────────────────────────────────────────────────

WEIGHT_TABLE = {
    6:  round(6**2  / 162, 3),
    8:  round(8**2  / 162, 3),
    10: round(10**2 / 162, 3),
    12: round(12**2 / 162, 3),
    16: round(16**2 / 162, 3),
    20: round(20**2 / 162, 3),
    25: round(25**2 / 162, 3),
    32: round(32**2 / 162, 3),
    40: round(40**2 / 162, 3),
}


def weight_per_meter(diameter_mm: int) -> float:
    """Return unit weight (kg/m) for a given rebar diameter using D²/162."""
    if diameter_mm in WEIGHT_TABLE:
        return WEIGHT_TABLE[diameter_mm]
    return round(diameter_mm**2 / 162, 3)


# ─────────────────────────────────────────────────────────────────────────────
# 2.  SAMPLE PDF GENERATOR
#     Creates a realistic rebar schedule PDF for demo / testing purposes.
# ─────────────────────────────────────────────────────────────────────────────

SAMPLE_DATA = [
    # member,        bar_mark, dia, length_m, qty, spacing_mm
    ("Footing F1",   "F1-01",  16,  3.20,     12,  150),
    ("Footing F1",   "F1-02",  16,  3.20,     12,  150),
    ("Footing F2",   "F2-01",  12,  2.80,      8,  200),
    ("Column C1",    "C1-01",  20,  3.00,      8,   None),
    ("Column C1",    "C1-02",   8,  1.10,     20,  180),
    ("Column C2",    "C2-01",  16,  3.00,      6,   None),
    ("Beam B1",      "B1-01",  16,  5.50,      4,   None),
    ("Beam B1",      "B1-02",  16,  5.50,      2,   None),
    ("Beam B1",      "B1-S1",   8,  1.20,     32,  150),
    ("Slab S1",      "S1-01",  10,  4.00,     25,  150),
    ("Slab S1",      "S1-02",  10,  3.50,     28,  150),
    ("Staircase ST", "ST-01",  12,  2.50,     18,  200),
]


def generate_sample_pdf(path: str = "sample_rebar_schedule.pdf"):
    """Generate a sample rebar schedule PDF that mimics a real BBS document."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(15, 15, 15)

    # ── Title block ──────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "BAR BENDING SCHEDULE (BBS)", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, "Project : Sample Residential Structure", ln=True)
    pdf.cell(0, 6, "Prepared by : Structural Engineer", ln=True)
    pdf.cell(0, 6, "Code Reference : IS 456 / D^2/162 formula for unit weight", ln=True)
    pdf.ln(4)

    # ── Weight reference table ───────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "Standard Unit Weights (kg/m) - Formula: D^2 / 162", ln=True)
    pdf.set_font("Helvetica", "", 9)
    header = "Dia(mm): " + "  |  ".join(
        [f"{d}mm = {w} kg/m" for d, w in list(WEIGHT_TABLE.items())[:6]]
    )
    pdf.multi_cell(0, 6, header)
    pdf.ln(3)

    # ── Main schedule table ──────────────────────────────────────────────────
    col_widths = [38, 18, 14, 22, 12, 20, 22, 22]
    headers    = ["Member", "Bar Mark", "Dia\n(mm)", "Length\n(m)", "Qty",
                  "Spacing\n(mm)", "Total\nLength (m)", "Total\nWeight (kg)"]

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(220, 220, 220)
    for w, h in zip(col_widths, headers):
        pdf.multi_cell(w, 5, h, border=1, align="C", fill=True, ln=3)
    pdf.ln(10)

    pdf.set_font("Helvetica", "", 8)
    for row in SAMPLE_DATA:
        member, mark, dia, length, qty, spacing = row
        total_length = round(length * qty, 2)
        total_weight = round(total_length * weight_per_meter(dia), 2)
        spacing_str  = str(spacing) if spacing else "N/A"
        values = [member, mark, str(dia), str(length), str(qty),
                  spacing_str, str(total_length), str(total_weight)]
        for w, v in zip(col_widths, values):
            pdf.cell(w, 6, v, border=1, align="C")
        pdf.ln()

    pdf.ln(6)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 6,
             "Note: Weight calculated using standard formula W = (D^2 / 162) x Length x Quantity",
             ln=True)

    pdf.output(path)
    print(f"[INFO] Sample PDF generated: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 3.  PDF TEXT EXTRACTION
#     Tries pdfplumber first (vector PDFs).
#     Falls back to Tesseract OCR for scanned/image PDFs.
# ─────────────────────────────────────────────────────────────────────────────

def extract_text_from_pdf(pdf_path: str) -> list[str]:
    """
    Extract raw text from every page.
    Returns a list of strings, one per page.
    """
    pages_text = []

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text and text.strip():
                pages_text.append(text)
                print(f"[INFO] Page {i+1}: extracted {len(text)} chars via pdfplumber (vector)")
            else:
                # Fallback: render page as image and run OCR
                print(f"[INFO] Page {i+1}: no selectable text, attempting OCR fallback...")
                text = _ocr_page(page)
                pages_text.append(text if text else "")

    return pages_text


def _ocr_page(page) -> str:
    """Render a pdfplumber page to image and run Tesseract OCR on it."""
    try:
        import pytesseract
        from PIL import Image
        import io

        img = page.to_image(resolution=300).original
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        pil_img = Image.open(buf)
        text = pytesseract.image_to_string(pil_img, config="--psm 6")
        print(f"[INFO]   OCR extracted {len(text)} chars via Tesseract")
        return text
    except ImportError:
        print("[WARN]  pytesseract / pillow not installed. Skipping OCR for this page.")
        print("        To enable: pip install pytesseract pillow  +  brew install tesseract")
        return ""
    except Exception as e:
        print(f"[WARN]  OCR failed: {e}")
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# 4.  TABLE EXTRACTION  (pdfplumber native tables)
# ─────────────────────────────────────────────────────────────────────────────

def extract_tables_from_pdf(pdf_path: str) -> list[list]:
    """
    Use pdfplumber's built-in table detector to pull structured tables.
    Returns a flat list of rows across all pages.
    """
    all_rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if any(cell and str(cell).strip() for cell in row):
                        all_rows.append(row)
    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# 5.  REBAR ROW PARSER
#     Identifies rows that look like rebar data using regex patterns.
# ─────────────────────────────────────────────────────────────────────────────

# Patterns to match rebar diameter annotations like T16, Y20, dia-12, 16mm, #16
DIA_PATTERN   = re.compile(r'\b(?:T|Y|R|N|dia[-\s]?)?(\d{1,2})\s*(?:mm)?\b')
# Patterns to match a length value like 3.20, 5500 (mm), 12m
LEN_PATTERN   = re.compile(r'\b(\d{1,2}\.?\d{0,3})\s*(?:m|mm)?\b')
# Patterns to match quantity (integer, typically 2-99 range for bar counts)
QTY_PATTERN   = re.compile(r'\b(\d{1,3})\b')
# Patterns to match spacing annotations like @150, c/c 200, 150c/c
SPC_PATTERN   = re.compile(r'[@c/\s]*(\d{2,4})\s*(?:mm|c[/.]?c)?', re.IGNORECASE)


VALID_DIAMETERS = {6, 8, 10, 12, 16, 20, 25, 28, 32, 40}


def parse_rebar_rows(raw_rows: list[list]) -> list[dict]:
    """
    Walk through extracted table rows and try to identify rebar data.
    Returns a list of parsed rebar records.
    """
    records = []

    for row in raw_rows:
        row_str = " | ".join(str(c) if c else "" for c in row)

        # Look for a valid rebar diameter in the row
        dia_matches = DIA_PATTERN.findall(row_str)
        dia = None
        for m in dia_matches:
            d = int(m)
            if d in VALID_DIAMETERS:
                dia = d
                break

        if dia is None:
            continue  # row does not contain a recognisable rebar diameter

        # Try to pull member name, bar mark, length, qty, spacing from cells
        cells = [str(c).strip() if c else "" for c in row]

        member   = cells[0] if len(cells) > 0 else "Unknown"
        bar_mark = cells[1] if len(cells) > 1 else ""

        # Length: look for a float in expected range (0.5m to 20m)
        length = None
        for cell in cells[2:]:
            nums = LEN_PATTERN.findall(cell)
            for n in nums:
                val = float(n)
                if 0.3 <= val <= 20.0:
                    length = val
                    break
            if length:
                break

        # Quantity: integer, reasonable bar count range 1-500
        qty = None
        for cell in cells:
            nums = QTY_PATTERN.findall(cell)
            for n in nums:
                val = int(n)
                if 1 <= val <= 500 and str(val) in cell:
                    qty = val
                    break
            if qty:
                break

        # Spacing (optional)
        spacing = None
        for cell in cells:
            m = SPC_PATTERN.search(cell)
            if m:
                s = int(m.group(1))
                if 50 <= s <= 600:
                    spacing = s
                    break

        if length is None or qty is None:
            continue  # not enough data to compute weight

        total_length = round(length * qty, 3)
        unit_wt      = weight_per_meter(dia)
        total_weight = round(total_length * unit_wt, 3)

        records.append({
            "Member"          : member,
            "Bar Mark"        : bar_mark,
            "Diameter (mm)"   : dia,
            "Length (m)"      : length,
            "Quantity"        : qty,
            "Spacing (mm)"    : spacing if spacing else "N/A",
            "Unit Weight(kg/m)": unit_wt,
            "Total Length (m)": total_length,
            "Total Weight (kg)": total_weight,
        })

    return records


# ─────────────────────────────────────────────────────────────────────────────
# 6.  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(records: list[dict], output_path: str = "rebar_output.xlsx"):
    """Write parsed rebar records to a formatted Excel file."""
    if not records:
        print("[WARN] No rebar records found to export.")
        return

    df = pd.DataFrame(records)

    # Summary row
    summary = {col: "" for col in df.columns}
    summary["Member"]            = "TOTAL"
    summary["Total Length (m)"]  = round(df["Total Length (m)"].sum(), 3)
    summary["Total Weight (kg)"] = round(df["Total Weight (kg)"].sum(), 3)
    df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Rebar Schedule", index=False)

        ws = writer.sheets["Rebar Schedule"]

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Highlight summary row
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = total_fill

    print(f"[INFO] Excel output saved: {output_path}")
    print(f"[INFO] Total rows extracted: {len(records)}")
    print(f"[INFO] Total rebar weight  : {df['Total Weight (kg)'].iloc[-1]} kg")


# ─────────────────────────────────────────────────────────────────────────────
# 7.  MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def run_extraction(pdf_path: str, output_path: str = None):
    """Full pipeline: PDF in, Excel out."""
    if not os.path.exists(pdf_path):
        sys.exit(f"[ERROR] File not found: {pdf_path}")

    if output_path is None:
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        output_path = f"{base}_rebar_output.xlsx"

    print(f"\n{'='*60}")
    print(f"  Rebar PDF Extractor")
    print(f"  Input  : {pdf_path}")
    print(f"  Output : {output_path}")
    print(f"{'='*60}\n")

    # Step 1: extract tables (structured path for vector PDFs)
    print("[STEP 1] Extracting tables from PDF...")
    raw_rows = extract_tables_from_pdf(pdf_path)
    print(f"[INFO]   {len(raw_rows)} raw table rows found\n")

    # Step 2: also extract plain text (for text-only layouts or OCR fallback)
    print("[STEP 2] Extracting page text...")
    pages_text = extract_text_from_pdf(pdf_path)

    # Step 3: parse rebar data from table rows
    print("\n[STEP 3] Parsing rebar records...")
    records = parse_rebar_rows(raw_rows)

    # Step 4: if table extraction yielded nothing, try text-line parsing
    if not records:
        print("[INFO]   Table parser found 0 records. Trying text-line fallback...")
        text_rows = []
        for page_text in pages_text:
            for line in page_text.split("\n"):
                cells = re.split(r"\s{2,}|\t|\|", line)
                if len(cells) >= 3:
                    text_rows.append(cells)
        records = parse_rebar_rows(text_rows)

    print(f"[INFO]   {len(records)} rebar records parsed\n")

    # Step 5: export to Excel
    print("[STEP 4] Exporting to Excel...")
    export_to_excel(records, output_path)

    print(f"\n[DONE] Check '{output_path}' for results.\n")


# ─────────────────────────────────────────────────────────────────────────────
# 8.  CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extract rebar schedule data from a PDF and export to Excel."
    )
    parser.add_argument(
        "--pdf",
        type=str,
        help="Path to the input PDF file containing the rebar schedule."
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Path for the output Excel file. Defaults to <input_name>_rebar_output.xlsx"
    )
    parser.add_argument(
        "--demo",
        action="store_true",
        help="Generate a sample rebar schedule PDF and run extraction on it (no input needed)."
    )

    args = parser.parse_args()

    if args.demo:
        sample_pdf = generate_sample_pdf("sample_rebar_schedule.pdf")
        run_extraction(sample_pdf, "sample_rebar_output.xlsx")
    elif args.pdf:
        run_extraction(args.pdf, args.output)
    else:
        parser.print_help()
        print("\n[TIP] Run with --demo to test immediately without any PDF:")
        print("      python rebar_pdf_extractor.py --demo\n")
